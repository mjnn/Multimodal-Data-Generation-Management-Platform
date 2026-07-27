#!/usr/bin/env python3
"""Parse OMS label taxonomy xlsx into YAML configuration."""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any

import yaml

SHEET_NAME = "DMS数据采集标签选择"
HEADER_ROW = 3
DATA_START_ROW = 4

_ARRAY_DTYPE_PATTERN = re.compile(
    r"^(bool|float|int|enum)\[(\d+)\](?:\s+per\s+seat)?$",
    re.IGNORECASE,
)


def load_config(config_path: Path) -> dict[str, Any]:
    with config_path.open(encoding="utf-8") as config_file:
        config = yaml.safe_load(config_file)
    if not isinstance(config, dict):
        raise ValueError(f"Invalid config file: {config_path}")
    return config


def split_enum_values(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [part.strip() for part in str(raw).split("/") if part.strip()]


def parse_value_schema(dtype: str, values: str | None) -> dict[str, Any]:
    normalized = dtype.strip()

    if normalized == "enum":
        return {"type": "enum", "values": split_enum_values(values)}

    if normalized == "bool":
        enum_values = split_enum_values(values)
        return {"type": "bool", "values": enum_values or ["true", "false"]}

    if normalized in {"int", "float", "string"}:
        schema: dict[str, Any] = {"type": normalized}
        if values:
            schema["range_hint"] = str(values)
        return schema

    if normalized == "enum[]":
        return {"type": "array", "items": {"type": "enum", "values": split_enum_values(values)}}

    if normalized == "list of enum":
        return {"type": "array", "items": {"type": "enum", "values": split_enum_values(values)}}

    if normalized == "enum + rationale":
        return {
            "type": "composite",
            "fields": [
                {"name": "value", "type": "enum"},
                {"name": "rationale", "type": "string"},
            ],
            "range_hint": str(values) if values else None,
        }

    if normalized == "int64 + string":
        return {
            "type": "composite",
            "fields": [
                {"name": "timestamp_ms", "type": "int64"},
                {"name": "timezone", "type": "string"},
            ],
            "range_hint": str(values) if values else None,
        }

    array_match = _ARRAY_DTYPE_PATTERN.match(normalized)
    if array_match:
        item_type = array_match.group(1).lower()
        length = int(array_match.group(2))
        schema = {"type": "array", "items": {"type": item_type}, "length": length}
        if "per seat" in normalized.lower():
            schema["per_seat"] = True
        if values:
            schema["range_hint"] = str(values)
        return schema

    raise ValueError(f"Unsupported dtype: {dtype}")


def parse_level(level_text: str) -> dict[str, str]:
    parts = level_text.strip().split(maxsplit=1)
    level_code = parts[0]
    level_name = parts[1] if len(parts) > 1 else ""
    return {"code": level_code, "name": level_name}


def read_labels_from_xlsx(xlsx_path: Path) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl is required to parse taxonomy xlsx") from exc

    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_NAME}")

    worksheet = workbook[SHEET_NAME]
    labels: list[dict[str, Any]] = []

    for row in worksheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if not row or not row[2]:
            continue

        level = parse_level(str(row[1]))
        label_id = str(row[2]).strip()
        labels.append(
            {
                "no": str(row[0]).strip() if row[0] is not None else "",
                "level_code": level["code"],
                "level_name": level["name"],
                "id": label_id,
                "name": str(row[3]).strip() if row[3] is not None else "",
                "definition": str(row[4]).strip() if row[4] is not None else "",
                "dtype": str(row[5]).strip() if row[5] is not None else "",
                "value_schema": parse_value_schema(str(row[5]), str(row[6]) if row[6] else None),
                "values_hint": str(row[6]).strip() if row[6] is not None else "",
                "selection_reason": str(row[7]).strip() if row[7] is not None else "",
            }
        )

    return labels


def build_taxonomy(
    labels: list[dict[str, Any]],
    *,
    source_file: str,
    exclude_labels: list[str] | None = None,
) -> dict[str, Any]:
    excluded = set(exclude_labels or [])
    active_labels = [label for label in labels if label["id"] not in excluded]

    return {
        "version": "v2",
        "source": source_file,
        "label_count": len(active_labels),
        "excluded_labels": sorted(excluded),
        "labels": active_labels,
    }


def write_taxonomy_yaml(taxonomy: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        yaml.safe_dump(taxonomy, allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )


def get_active_label_ids(taxonomy: dict[str, Any]) -> list[str]:
    return [str(label["id"]) for label in taxonomy.get("labels", [])]


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate OMS label taxonomy YAML from xlsx.")
    parser.add_argument(
        "--config",
        type=Path,
        default=Path("config.yaml"),
        help="Path to config file (default: ./config.yaml)",
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        help="Path to taxonomy xlsx (default: project root DMS file)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output YAML path (default: cloud.job3_label.label_taxonomy)",
    )
    parser.add_argument(
        "--exclude",
        action="append",
        dest="exclude_labels",
        default=[],
        help="Label id to exclude; can be repeated.",
    )
    args = parser.parse_args()

    config_path = args.config.resolve()
    config = load_config(config_path)
    project_root = config_path.parent

    job3_config = config.get("cloud", {}).get("job3_label", {})
    xlsx_path = args.xlsx or (project_root / "DMS数据采集标签选择方案_v2.xlsx")
    output_path = args.output or (project_root / str(job3_config.get("label_taxonomy", "config/oms_label_taxonomy.yaml")))

    config_excludes = job3_config.get("exclude_labels", [])
    exclude_labels = list(dict.fromkeys([*config_excludes, *args.exclude_labels]))

    labels = read_labels_from_xlsx(xlsx_path.resolve())
    taxonomy = build_taxonomy(
        labels,
        source_file=str(xlsx_path.name),
        exclude_labels=exclude_labels,
    )
    write_taxonomy_yaml(taxonomy, output_path.resolve())

    print(f"Wrote {taxonomy['label_count']} label(s) to {output_path}")
    if exclude_labels:
        print(f"Excluded {len(exclude_labels)} label(s): {', '.join(exclude_labels)}")


if __name__ == "__main__":
    main()
